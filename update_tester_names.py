import openpyxl

file_path = r"C:\Users\Admin\Downloads\Testing Report Format - Detailed.xlsx"
wb = openpyxl.load_workbook(file_path)
ws_tc = wb.worksheets[0]

# Counters for alternating
ngo_admin_testers = ["Ayush Potdar", "Sarthak Pawar"]
ngo_admin_idx = 0

super_admin_testers = ["Owais Khan", "Shivraj Patil"]
super_admin_idx = 0

# Tester columns (Assuming Sr.No=A(1), ..., Tester Name=K(11))
TESTER_COL = 11

def determine_tester(row_values):
    global ngo_admin_idx, super_admin_idx
    
    # row_values: [Sr.No, Sc_Id, Sc_Obj, Tc_Id, Tc_Obj, Steps, Data, Exp, Act, Status, Tester, Remark]
    text_content = str(row_values[2]).lower() + " " + str(row_values[4]).lower() + " " + str(row_values[5]).lower()
    
    # We use a point system or keyword priority
    
    # 1. Super Admin
    if "super admin" in text_content or "user management" in text_content or "system settings" in text_content or "activity logs" in text_content or "global notifications" in text_content:
        tester = super_admin_testers[super_admin_idx]
        super_admin_idx = (super_admin_idx + 1) % 2
        return tester
        
    # 2. NGO Admin
    if "ngo admin" in text_content or "campaign management" in text_content or "ngo dashboard" in text_content or "contact inquiry" in text_content or "ngo profile" in text_content or "create campaign" in text_content or "ngo management" in text_content:
        tester = ngo_admin_testers[ngo_admin_idx]
        ngo_admin_idx = (ngo_admin_idx + 1) % 2
        return tester
        
    # 3. Donor
    if "donor" in text_content or "donation" in text_content or "razorpay" in text_content or "receipt" in text_content or "payment" in text_content:
        return "Radhika Panchal"
        
    # 4. Event Coordinator
    if "coordinator" in text_content or "event creation" in text_content or "task assignment" in text_content or "assign task" in text_content or "revise task" in text_content or "revision" in text_content:
        return "Prathamesh Shimpale"
        
    # 5. Volunteer
    if "volunteer" in text_content or "my events" in text_content or "proof upload" in text_content or "submit proof" in text_content or "my tasks" in text_content or "tree plantation" in text_content:
        return "Pratik Rahane"
        
    # 6. Authentication
    if "login" in text_content or "logout" in text_content or "password" in text_content or "session" in text_content or "credential" in text_content or "auth" in text_content:
        return "Roshni Patole"
        
    # 7. Landing Page
    if "landing page" in text_content or "contact us" in text_content or "ui responsiveness" in text_content or "responsive" in text_content or "mobile" in text_content or "home page" in text_content:
        return "Roshni Patole"
        
    # Fallback to Super Admin or NGO Admin if reports/analytics/tables
    if "report" in text_content or "analytic" in text_content or "grid" in text_content or "table" in text_content or "pagination" in text_content:
        tester = super_admin_testers[super_admin_idx]
        super_admin_idx = (super_admin_idx + 1) % 2
        return tester
        
    # Security / XSS fallback
    if "sql injection" in text_content or "xss" in text_content or "security" in text_content:
        return "Roshni Patole"

    return "Roshni Patole" # default fallback

for row in ws_tc.iter_rows(min_row=5, max_row=ws_tc.max_row):
    if not row[0].value:
        continue # empty row
        
    row_values = [cell.value for cell in row]
    tester = determine_tester(row_values)
    
    # Column K is index 10 (0-based) in row tuple
    row[10].value = tester

new_file_path = r"C:\Users\Admin\Downloads\Testing Report - Final.xlsx"
wb.save(new_file_path)
print("Tester names successfully updated.")
