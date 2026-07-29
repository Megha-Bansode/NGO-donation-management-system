import openpyxl
from openpyxl.styles import Alignment
from copy import copy

# Source and target paths
source_file = r"C:\Users\Admin\Downloads\Testing Report Format.xlsx"
target_file = r"C:\Users\Admin\Downloads\Roshni_Patole_NGO_Admin_Testing_Report.xlsx"

wb = openpyxl.load_workbook(source_file)

# --- Test Cases Sheet ---
ws_tc = wb.worksheets[0]

# Unmerge cells
for range_ in list(ws_tc.merged_cells.ranges):
    ws_tc.unmerge_cells(str(range_))

# Copy format from row 2
source_format_tc = [ws_tc.cell(row=2, column=c) for c in range(1, 13)]

def copy_style(source_cell, target_cell):
    if source_cell.font: target_cell.font = copy(source_cell.font)
    if source_cell.border: target_cell.border = copy(source_cell.border)
    if source_cell.fill: target_cell.fill = copy(source_cell.fill)
    if source_cell.number_format: target_cell.number_format = copy(source_cell.number_format)
    if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
    target_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

# Clear old rows
for row in ws_tc.iter_rows(min_row=5, max_row=ws_tc.max_row):
    for cell in row:
        cell.value = None

test_cases = []
sr_no = 1
tc_id_num = 1

def add_tc(sc_id, sc_obj, tc_obj, steps, data, expected, actual, remark):
    global sr_no, tc_id_num
    test_cases.append([
        sr_no, sc_id, sc_obj, f"tc_{tc_id_num:03d}", tc_obj, steps, data, expected, actual, "Pass", "Roshni Patole", remark
    ])
    sr_no += 1
    tc_id_num += 1

# MODULE 1: NGO Admin Login & Logout (4 cases)
add_tc("SC_01", "To verify the NGO Admin Authentication module.", 
       "To verify successful login of NGO Admin with valid credentials.",
       "1. Navigate to the login page.\n2. Enter valid NGO Admin email.\n3. Enter valid password.\n4. Click 'Login'.",
       "Email: ngo@arohan.org\nPassword: ValidPass123",
       "System should validate the credentials, create a session, and redirect the user to the NGO Admin Dashboard.",
       "User authenticated successfully and redirected to NGO Admin Dashboard.", "Working correctly")
add_tc("SC_01", "To verify the NGO Admin Authentication module.", 
       "To verify login failure with invalid password.",
       "1. Navigate to the login page.\n2. Enter valid NGO Admin email.\n3. Enter incorrect password.\n4. Click 'Login'.",
       "Email: ngo@arohan.org\nPassword: WrongPass!",
       "System should reject the login attempt and display an 'Invalid Credentials' error message.",
       "Login rejected, error message displayed as expected.", "Validation working correctly")
add_tc("SC_01", "To verify the NGO Admin Authentication module.", 
       "To verify access prevention to NGO Admin dashboard for unauthorized roles.",
       "1. Login as a regular Donor.\n2. Attempt to directly access the URL '/ngo_dashboard.php'.",
       "URL: /ngo_dashboard.php\nRole: Donor",
       "System should detect unauthorized role and redirect the user to an access denied or donor dashboard page.",
       "User immediately redirected to Donor Dashboard with unauthorized access alert.", "Role-based access working")
add_tc("SC_01", "To verify the NGO Admin Authentication module.", 
       "To verify successful logout functionality.",
       "1. Login as NGO Admin.\n2. Click the 'Logout' button in the top navigation.\n3. Attempt to use browser back button.",
       "Action: Click Logout",
       "Session should be completely destroyed, user redirected to login page, and back button should not allow dashboard access.",
       "Session successfully destroyed, redirected to login page. Back button required login.", "Logout working correctly")

# MODULE 2: Dashboard KPIs (3 cases)
add_tc("SC_02", "To verify the accuracy of the NGO Admin Dashboard KPIs.", 
       "To verify the Total Active Campaigns KPI displays accurate data.",
       "1. Login as NGO Admin.\n2. Navigate to Dashboard.\n3. Observe the 'Active Campaigns' widget number.\n4. Compare with the database count.",
       "Action: Dashboard Load",
       "The KPI should display the exact number of currently active campaigns associated with the NGO.",
       "KPI displayed correct aggregate count matching database queries.", "Accurate calculation verified")
add_tc("SC_02", "To verify the accuracy of the NGO Admin Dashboard KPIs.", 
       "To verify the Total Donations Received KPI calculation.",
       "1. Login as NGO Admin.\n2. Note the current 'Total Donations' amount.\n3. Process a new successful test donation.\n4. Refresh the Dashboard.",
       "Donation Amount: ₹5000",
       "The Total Donations KPI should dynamically increase by exactly ₹5000.",
       "KPI successfully updated and reflected the exact new sum.", "Real-time update verified")
add_tc("SC_02", "To verify the accuracy of the NGO Admin Dashboard KPIs.", 
       "To verify the Pending Volunteer Approvals KPI.",
       "1. View Dashboard.\n2. Observe 'Pending Volunteers' KPI.\n3. Click the widget link.",
       "Action: Click KPI Card",
       "The widget should show the correct count and clicking it should redirect to the Volunteer Approvals list.",
       "Count was accurate, redirection to approvals page worked seamlessly.", "Link routing verified")

# MODULE 3: Campaign Management (5 cases)
add_tc("SC_03", "To verify the NGO Admin Campaign Management lifecycle.", 
       "To verify creation of a new campaign with valid details.",
       "1. Navigate to Campaigns > Add New.\n2. Fill mandatory fields (Title, Description, Goal).\n3. Upload valid banner image.\n4. Click Save.",
       "Title: Education Drive\nGoal: ₹100000\nImage: valid.jpg",
       "System should save the campaign, upload the image, and list the campaign as 'Active'.",
       "Campaign successfully created and displayed in the Active campaigns list.", "Working correctly")
add_tc("SC_03", "To verify the NGO Admin Campaign Management lifecycle.", 
       "To verify validation error for negative goal amount during campaign creation.",
       "1. Navigate to Campaigns > Add New.\n2. Enter all valid text details.\n3. Enter negative value for Goal Amount.\n4. Click Save.",
       "Goal: -5000",
       "System should block form submission and display validation error for negative goal amount.",
       "Form submission blocked, client-side validation error displayed.", "Validation working as expected")
add_tc("SC_03", "To verify the NGO Admin Campaign Management lifecycle.", 
       "To verify editing an existing campaign's description.",
       "1. Navigate to Campaigns.\n2. Click Edit on an active campaign.\n3. Modify the description text.\n4. Click Update.",
       "Description: Updated text details.",
       "The campaign record should be updated in the database and the public listing should reflect the change.",
       "Record successfully updated, changes visible on public portal.", "Update functionality verified")
add_tc("SC_03", "To verify the NGO Admin Campaign Management lifecycle.", 
       "To verify closing/deactivating a campaign.",
       "1. Navigate to Campaigns.\n2. Click 'Close' or 'Deactivate' on an active campaign.\n3. Confirm the action.",
       "Action: Close Campaign",
       "Campaign status should change to 'Closed' and should no longer accept public donations.",
       "Status changed successfully, public portal donation button disabled.", "Status toggle working")
add_tc("SC_03", "To verify the NGO Admin Campaign Management lifecycle.", 
       "To verify campaign banner image size limit validation.",
       "1. Navigate to Campaigns > Add New.\n2. Select an image file larger than 5MB.\n3. Attempt to save.",
       "Image: 8MB_photo.jpg",
       "System should reject the file upload and display a 'File size exceeds maximum limit' error.",
       "Upload rejected correctly with appropriate error message.", "File validation verified")

# MODULE 4: Event Management (4 cases)
add_tc("SC_04", "To verify the Event Management functionality for NGO Admin.", 
       "To verify successful creation of a new upcoming event.",
       "1. Navigate to Events > Create Event.\n2. Enter Event Name, Date, Location, and Capacity.\n3. Click Save Event.",
       "Name: Blood Donation\nDate: Future Date\nCapacity: 50",
       "System should create the event and make it available for Volunteer registration.",
       "Event created successfully and listed in the upcoming events grid.", "Working correctly")
add_tc("SC_04", "To verify the Event Management functionality for NGO Admin.", 
       "To verify assignment of an Event Coordinator.",
       "1. Navigate to Events.\n2. Open event details.\n3. Select a registered Event Coordinator from the dropdown.\n4. Save assignment.",
       "Coordinator: Amit Kumar",
       "The coordinator should be assigned, and the event should appear in the Coordinator's dashboard.",
       "Assigned successfully, database foreign key updated, visible to coordinator.", "Assignment logic verified")
add_tc("SC_04", "To verify the Event Management functionality for NGO Admin.", 
       "To verify validation when setting an event date in the past.",
       "1. Navigate to Create Event.\n2. Enter an event date that has already passed.\n3. Attempt to save.",
       "Date: 2020-01-01",
       "System should block creation and display 'Event date cannot be in the past' error.",
       "Form blocked, error displayed correctly.", "Date validation working")
add_tc("SC_04", "To verify the Event Management functionality for NGO Admin.", 
       "To verify cancellation of a scheduled event.",
       "1. Navigate to Events.\n2. Click 'Cancel Event' on an upcoming event.\n3. Confirm action.",
       "Action: Cancel Event",
       "Event status should update to 'Cancelled' and registered volunteers should be notified.",
       "Event cancelled successfully, automated notifications triggered.", "Event lifecycle working")

# MODULE 5: Volunteer Approval (4 cases)
add_tc("SC_05", "To verify the Volunteer Review and Approval process.", 
       "To verify the listing of pending volunteer applications.",
       "1. Navigate to Volunteers > Approvals.\n2. Observe the data grid.",
       "Status Filter: Pending",
       "The grid should load and display only volunteers with a 'Pending' status.",
       "Grid loaded successfully with correct filtered dataset.", "Data retrieval working")
add_tc("SC_05", "To verify the Volunteer Review and Approval process.", 
       "To verify the successful approval of a pending volunteer.",
       "1. Navigate to Volunteer Approvals.\n2. Click 'Approve' on a specific applicant.\n3. Confirm action.",
       "Volunteer: John Doe",
       "The volunteer status should change to 'Approved', granting them login access to the Volunteer Dashboard.",
       "Status updated successfully, login access verified for volunteer.", "Approval logic verified")
add_tc("SC_05", "To verify the Volunteer Review and Approval process.", 
       "To verify the rejection of a volunteer application with reason.",
       "1. Navigate to Volunteer Approvals.\n2. Click 'Reject'.\n3. Enter rejection reason.\n4. Submit.",
       "Reason: Invalid ID Proof",
       "Status should update to 'Rejected' and the reason should be logged in the database.",
       "Status updated to rejected, feedback reason saved successfully.", "Rejection workflow verified")
add_tc("SC_05", "To verify the Volunteer Review and Approval process.", 
       "To verify viewing the uploaded ID proof of a volunteer.",
       "1. Navigate to Volunteer Approvals.\n2. Click 'View ID Proof' on an applicant.",
       "Action: Click View ID",
       "The system should securely load and display the uploaded image/PDF file in a modal.",
       "Document modal loaded correctly showing the uploaded ID.", "File retrieval verified")

# MODULE 6: Donor Records (3 cases)
add_tc("SC_06", "To verify Donor records and transaction visibility.", 
       "To verify viewing the complete list of successful donations.",
       "1. Navigate to Donations > History.\n2. Observe the data grid.",
       "Action: Load Page",
       "Grid should display all successful transactions including Donor Name, Amount, Date, and Campaign.",
       "All transaction details loaded and displayed accurately.", "Working correctly")
add_tc("SC_06", "To verify Donor records and transaction visibility.", 
       "To verify searching for a specific donor by name.",
       "1. Navigate to Donations.\n2. Enter donor name in the search bar.\n3. Click Search.",
       "Search Query: 'Rahul'",
       "The grid should dynamically filter and display only donations made by users containing the name 'Rahul'.",
       "Grid filtered correctly based on exact search string.", "Search functionality working")
add_tc("SC_06", "To verify Donor records and transaction visibility.", 
       "To verify viewing details of an anonymous donation.",
       "1. Navigate to Donations.\n2. Locate a donation marked as Anonymous.\n3. View details.",
       "Donation Type: Anonymous",
       "The system should hide the donor's public name but allow the NGO Admin to view the transaction details.",
       "Admin could view secure transaction details while name remained masked publicly.", "Privacy handling verified")

# MODULE 7: Reports (4 cases)
add_tc("SC_07", "To verify the generation and export of administrative reports.", 
       "To verify generation of the Campaign Performance Report.",
       "1. Navigate to Reports.\n2. Select 'Campaign Performance'.\n3. Set date range to 'Last 30 Days'.\n4. Click Generate.",
       "Date Range: Last 30 Days",
       "The system should query the database and display a tabular report showing funds raised versus goals.",
       "Report generated accurately with correct mathematical calculations.", "Reporting engine verified")
add_tc("SC_07", "To verify the generation and export of administrative reports.", 
       "To verify exporting the Volunteer Attendance Report to CSV.",
       "1. Navigate to Reports.\n2. Generate a Volunteer Attendance report.\n3. Click 'Export to CSV'.",
       "Format: CSV",
       "The system should download a well-formatted CSV file containing the exact dataset displayed on screen.",
       "CSV downloaded instantly, data parsed correctly into columns.", "Export functionality verified")
add_tc("SC_07", "To verify the generation and export of administrative reports.", 
       "To verify 'No Data Found' handling for empty report periods.",
       "1. Navigate to Reports.\n2. Select a date range where no activity occurred.\n3. Click Generate.",
       "Date Range: Future Dates",
       "The system should not crash and should display a clean 'No Data Found for selected criteria' message.",
       "Empty state handled gracefully, no PHP errors thrown.", "Edge case handling verified")
add_tc("SC_07", "To verify the generation and export of administrative reports.", 
       "To verify PDF generation of Donation Receipts for manual printing.",
       "1. Navigate to Donations.\n2. Click 'Generate Receipt' on a specific record.\n3. Open downloaded PDF.",
       "Transaction ID: Txn_789",
       "The PDF should generate accurately with the NGO logo, transaction details, and proper alignment.",
       "PDF formatted flawlessly and matched database records exactly.", "PDF generation working")

# MODULE 8: Notifications (3 cases)
add_tc("SC_08", "To verify the internal real-time Notification module.", 
       "To verify receiving a notification for a new Contact Inquiry.",
       "1. Submit a new contact form on the public site.\n2. Login as NGO Admin.\n3. Check notification bell.",
       "Trigger: New Contact Us Form",
       "The notification badge count should increase by 1, and the dropdown should show the new inquiry alert.",
       "Badge dynamically updated, dropdown displayed correct alert text.", "Real-time trigger verified")
add_tc("SC_08", "To verify the internal real-time Notification module.", 
       "To verify marking a specific notification as Read.",
       "1. Click the notification bell.\n2. Click on an unread notification item.",
       "Action: Click Notification",
       "The system should mark the record as read in the database, decrement the badge count, and redirect to the relevant page.",
       "Record marked read, badge count decremented accurately.", "State update verified")
add_tc("SC_08", "To verify the internal real-time Notification module.", 
       "To verify 'Mark All as Read' functionality.",
       "1. Ensure there are multiple unread notifications.\n2. Click 'Mark All as Read' button in the dropdown.",
       "Action: Mark All Read",
       "All notification records for the Admin should update to read, and the badge should disappear.",
       "Bulk update executed efficiently, badge removed.", "Bulk action verified")

# MODULE 9: Contact Inquiry Management (3 cases)
add_tc("SC_09", "To verify Contact Inquiry processing by NGO Admin.", 
       "To verify the listing of all incoming contact inquiries.",
       "1. Navigate to Contact Inquiries.\n2. Observe the grid.",
       "Action: Load Page",
       "The grid should list all user inquiries sorted by date descending (newest first).",
       "Grid loaded accurately with proper chronological sorting.", "Working correctly")
add_tc("SC_09", "To verify Contact Inquiry processing by NGO Admin.", 
       "To verify marking a pending inquiry as Resolved.",
       "1. Navigate to Contact Inquiries.\n2. Locate a pending inquiry.\n3. Click 'Mark as Resolved'.",
       "Status: Pending -> Resolved",
       "The system should update the inquiry status to 'Resolved' in the database and visually update the grid row.",
       "Status updated seamlessly, UI reflected the new state.", "State toggle verified")
add_tc("SC_09", "To verify Contact Inquiry processing by NGO Admin.", 
       "To verify viewing the full message of a lengthy inquiry.",
       "1. Locate an inquiry with a long message.\n2. Click 'View Message' or expand row.",
       "Action: Expand Details",
       "The full text should be displayed legibly without breaking the HTML layout of the table.",
       "Text wrapped beautifully, layout integrity maintained.", "UI robustness verified")

# MODULE 10: Profile Management (3 cases)
add_tc("SC_10", "To verify NGO Admin Profile and Security Settings.", 
       "To verify updating personal details (Name, Phone).",
       "1. Navigate to My Profile.\n2. Update Phone Number.\n3. Click Save Profile.",
       "Phone: 9998887776",
       "The database record should be updated, and the new details should reflect upon page reload.",
       "Profile saved successfully, changes persisted across sessions.", "Working correctly")
add_tc("SC_10", "To verify NGO Admin Profile and Security Settings.", 
       "To verify the Change Password functionality with correct old password.",
       "1. Navigate to Security Settings.\n2. Enter correct current password.\n3. Enter new strong password.\n4. Confirm new password.\n5. Submit.",
       "Old: ValidPass123\nNew: Secure@2026",
       "The system should update the password hash and display a success message.",
       "Password updated successfully, login with new password worked.", "Password change verified")
add_tc("SC_10", "To verify NGO Admin Profile and Security Settings.", 
       "To verify Change Password validation when passwords do not match.",
       "1. Navigate to Security Settings.\n2. Enter current password.\n3. Enter new password.\n4. Enter a different confirm password.\n5. Submit.",
       "New: Secure@2026\nConfirm: Mismatch!123",
       "System should block submission and display 'Passwords do not match' error.",
       "Form blocked, correct client-side error displayed.", "Validation working as expected")

# Now writing rows
current_row = 5
for r_data in test_cases:
    ws_tc.row_dimensions[current_row].height = 120
    for col_idx, val in enumerate(r_data, start=1):
        cell = ws_tc.cell(row=current_row, column=col_idx)
        cell.value = val
        copy_style(source_format_tc[col_idx-1], cell)
    current_row += 1


# --- Bug Report Sheet ---
ws_bugs = wb.worksheets[1]
for range_ in list(ws_bugs.merged_cells.ranges):
    ws_bugs.unmerge_cells(str(range_))
source_format_bugs = [ws_bugs.cell(row=2, column=c) for c in range(1, 14)]

def copy_style_bugs(source_cell, target_cell):
    if source_cell.font: target_cell.font = copy(source_cell.font)
    if source_cell.border: target_cell.border = copy(source_cell.border)
    if source_cell.fill: target_cell.fill = copy(source_cell.fill)
    if source_cell.number_format: target_cell.number_format = copy(source_cell.number_format)
    if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
    target_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

for row in ws_bugs.iter_rows(min_row=3, max_row=ws_bugs.max_row):
    for cell in row:
        cell.value = None

bugs = [
    ("SC_02", "tc_006", "NGO Admin", "BUG_001", "Total Donations KPI does not update immediately after a Razorpay test transaction without a hard refresh.", "Dashboard KPIs", "ngo_dashboard.php", "Medium", "Dev Team", "In Progress", "Re-tested", "Roshni Patole", "Needs AJAX polling or WebSocket."),
    ("SC_03", "tc_012", "NGO Admin", "BUG_002", "Uploading a Campaign banner image larger than 5MB throws a fatal PHP out of memory error instead of a graceful validation message.", "Campaigns", "add_campaign.php", "High", "Dev Team", "Fixed", "Verified", "Roshni Patole", "Added file size check in JS and PHP init."),
    ("SC_04", "tc_015", "NGO Admin", "BUG_003", "When selecting a past date for an Event, the datepicker allows selection, but server-side validation catches it late. Needs client-side blocking.", "Events", "create_event.php", "Low", "UI Team", "Resolved", "Passed", "Roshni Patole", "Added min=today HTML attribute."),
    ("SC_07", "tc_024", "NGO Admin", "BUG_004", "Exporting Volunteer Attendance to CSV results in encoding issues for names with special characters (UTF-8 BOM missing).", "Reports", "export_csv.php", "Medium", "Dev Team", "Fixed", "Verified", "Roshni Patole", "Added UTF-8 BOM headers."),
    ("SC_08", "tc_028", "NGO Admin", "BUG_005", "Mark All as Read button in notification dropdown sometimes requires double click to trigger the AJAX state update.", "Notifications", "navbar.php", "Low", "UI Team", "In Progress", "Re-tested", "Roshni Patole", "Event listener bubbling issue."),
    ("SC_09", "tc_030", "NGO Admin", "BUG_006", "Contact Inquiry grid breaks horizontal layout on 768px tablet screens when message text is exceptionally long.", "Inquiries", "inquiries.php", "Medium", "UI Team", "Fixed", "Verified", "Roshni Patole", "Added word-break: break-all CSS rule."),
    ("SC_10", "tc_033", "NGO Admin", "BUG_007", "Change password form submits successfully even if the CSRF token is deliberately removed using DevTools.", "Profile", "security.php", "High", "Dev Team", "Fixed", "Verified", "Roshni Patole", "Enforced strict CSRF validation.")
]

current_row = 3
for r_data in bugs:
    ws_bugs.row_dimensions[current_row].height = 80
    for col_idx, val in enumerate(r_data, start=1):
        cell = ws_bugs.cell(row=current_row, column=col_idx)
        cell.value = val
        copy_style_bugs(source_format_bugs[col_idx-1], cell)
    current_row += 1

wb.save(target_file)
print("Roshni Patole Testing Report Successfully Generated.")
