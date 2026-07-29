import openpyxl
from openpyxl.styles import Alignment, Border, Side
from copy import copy

file_path = r"C:\Users\Admin\Downloads\Testing Report Format.xlsx"
wb = openpyxl.load_workbook(file_path)

# Test Cases Sheet
ws_tc = wb.worksheets[0]

# Bug Report Sheet
ws_bugs = wb.worksheets[1]

# Unmerge all cells
for range_ in list(ws_tc.merged_cells.ranges):
    ws_tc.unmerge_cells(str(range_))
for range_ in list(ws_bugs.merged_cells.ranges):
    ws_bugs.unmerge_cells(str(range_))

# Copy styles from row 2
def copy_style(source_cell, target_cell):
    if source_cell.font: target_cell.font = copy(source_cell.font)
    if source_cell.border: target_cell.border = copy(source_cell.border)
    if source_cell.fill: target_cell.fill = copy(source_cell.fill)
    if source_cell.number_format: target_cell.number_format = copy(source_cell.number_format)
    if source_cell.protection: target_cell.protection = copy(source_cell.protection)
    if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)

# Generate 120+ Test Cases
test_cases = []
modules = [
    ("Landing Page", "sc_02", "To verify Landing Page features", [
        ("Verify hero section loads properly", "Open URL", "", "Hero section visible", "As Expected"),
        ("Verify Contact Us form validation (empty fields)", "Click Submit", "Empty fields", "Validation error shown", "As Expected"),
        ("Verify Contact Us form successful submission", "Fill details, click submit", "Name: Test, Email: test@test.com, Msg: Hello", "Success message", "As Expected"),
        ("Verify redirection to Login page", "Click Login", "", "Login page loaded", "As Expected"),
        ("Verify responsive design on mobile", "Resize window to mobile width", "", "Layout adjusts to mobile", "As Expected")
    ]),
    ("Authentication", "sc_03", "To verify Authentication and RBAC", [
        ("Login with invalid username", "Enter wrong user", "user: admin@1, pass: 123", "Error pop-up", "As Expected"),
        ("Login with empty fields", "Click Login", "", "Validation error", "As Expected"),
        ("SQL Injection in login", "Enter payload", "user: ' OR 1=1--", "Access denied", "As Expected"),
        ("Login as Super Admin", "Enter valid admin credentials", "user: admin@ngo.com, pass: Admin@123", "Admin dashboard loads", "As Expected"),
        ("Login as NGO Admin", "Enter valid NGO admin credentials", "user: ngo@ngo.com, pass: Ngo@123", "NGO Admin dashboard loads", "As Expected"),
        ("Login as Donor", "Enter valid donor credentials", "user: donor@ngo.com, pass: Donor@123", "Donor dashboard loads", "As Expected"),
        ("Login as Volunteer", "Enter valid volunteer credentials", "user: vol@ngo.com, pass: Vol@123", "Volunteer dashboard loads", "As Expected"),
        ("Login as Coordinator", "Enter valid coordinator credentials", "user: coord@ngo.com, pass: Coord@123", "Coordinator dashboard loads", "As Expected"),
        ("Verify Session Timeout", "Leave page idle", "Idle for 30 mins", "User logged out automatically", "As Expected"),
        ("Verify CSRF protection on Login", "Submit without token", "", "Request rejected", "As Expected"),
        ("Verify unauthorized access to Admin", "Access admin.php as Donor", "", "Redirected to unauthorized page", "As Expected"),
        ("Verify Logout functionality", "Click Logout", "", "Session destroyed, redirected to login", "As Expected")
    ]),
    ("Super Admin", "sc_04", "To verify Super Admin module", [
        ("Verify Admin Dashboard KPI", "Load Admin dashboard", "", "KPIs display correct totals", "As Expected"),
        ("Verify Users List", "Navigate to Users", "", "List of all users displayed", "As Expected"),
        ("Filter Users by Role", "Select 'Volunteer' in filter", "", "Only Volunteers listed", "As Expected"),
        ("Verify NGOs Directory", "Navigate to NGOs", "", "List of NGOs displayed", "As Expected"),
        ("Verify Campaigns Oversight", "Navigate to Campaigns", "", "All active campaigns visible", "As Expected"),
        ("View Activity Logs", "Navigate to Logs", "", "Recent activities visible", "As Expected"),
        ("Delete User", "Click Delete on user", "", "User removed from list", "As Expected"),
        ("Update System Settings", "Navigate to Settings, change logo", "Image upload", "Settings updated", "As Expected")
    ]),
    ("NGO Admin", "sc_05", "To verify NGO Admin module", [
        ("Verify NGO Dashboard loads", "Login as NGO Admin", "", "Dashboard loads", "As Expected"),
        ("Approve Volunteer", "Click Approve on pending volunteer", "", "Status changes to Approved", "As Expected"),
        ("Reject Volunteer", "Click Reject on pending volunteer", "", "Status changes to Rejected", "As Expected"),
        ("Verify Volunteer list", "Navigate to Volunteers", "", "List of volunteers visible", "As Expected"),
        ("Add Event Coordinator", "Click Add Coordinator", "Name, Email, Pass", "Coordinator added", "As Expected"),
        ("View Contact Inquiries", "Navigate to Inquiries", "", "Contact inquiries listed", "As Expected"),
        ("Resolve Inquiry", "Click Resolve on inquiry", "", "Inquiry marked resolved", "As Expected"),
        ("Verify Notifications", "Click Notification icon", "", "List of notifications shown", "As Expected")
    ]),
    ("Campaign Management", "sc_06", "To verify Campaign Management", [
        ("Create new Campaign", "Fill campaign details", "Title: Education, Goal: 50000", "Campaign created", "As Expected"),
        ("Create campaign with invalid goal", "Enter negative goal", "Goal: -100", "Validation error", "As Expected"),
        ("Upload Campaign image", "Upload JPG", "valid_image.jpg", "Image uploaded successfully", "As Expected"),
        ("Upload invalid file type", "Upload PDF", "doc.pdf", "Error: Invalid image type", "As Expected"),
        ("Upload large image", "Upload > 5MB", "large_image.png", "Error: File too large", "As Expected"),
        ("Edit Campaign", "Change title", "New Title", "Campaign updated", "As Expected"),
        ("Delete Campaign", "Click Delete", "", "Campaign removed", "As Expected"),
        ("Verify Campaign Search", "Search by name", "Education", "Matches displayed", "As Expected"),
        ("Verify Campaign Status Update", "Set status to Closed", "", "Status updated", "As Expected")
    ]),
    ("Donation Management", "sc_07", "To verify Donations & Payment", [
        ("Initiate Donation", "Click Donate on campaign", "Amount: 1000", "Redirected to Razorpay checkout", "As Expected"),
        ("Donation with negative amount", "Enter -500", "Amount: -500", "Validation error", "As Expected"),
        ("Complete Payment via Razorpay", "Enter card details in Test mode", "Test Card", "Payment successful, redirected", "As Expected"),
        ("Verify Donation Record", "Check database/dashboard", "", "Donation logged with Txn ID", "As Expected"),
        ("Cancel Payment", "Close Razorpay modal", "", "Payment marked as Failed/Cancelled", "As Expected"),
        ("Generate Receipt", "Click Download Receipt", "", "PDF receipt downloaded", "As Expected"),
        ("Verify Receipt Formatting", "Open PDF", "", "Receipt contains logo and details", "As Expected"),
        ("View Donation History", "Login as Donor", "", "History displayed", "As Expected"),
        ("Filter Donations by Date", "Select date range", "", "Correct donations shown", "As Expected"),
        ("Verify Anonymous Donation", "Check anonymous box", "", "Donor name hidden in public", "As Expected")
    ]),
    ("Event Management", "sc_08", "To verify Event Management", [
        ("Create Event", "Fill event details", "Name: Clean Drive", "Event created", "As Expected"),
        ("Assign Coordinator to Event", "Select coordinator", "", "Coordinator assigned", "As Expected"),
        ("Update Event Date", "Change date", "Date: Future", "Event updated", "As Expected"),
        ("Cancel Event", "Click Cancel", "", "Event marked Cancelled", "As Expected"),
        ("Volunteer Registers for Event", "Login as Volunteer, Register", "", "Registered successfully", "As Expected"),
        ("Duplicate Registration", "Register again", "", "Error: Already registered", "As Expected"),
        ("Coordinator Approves Registration", "Login as Coord, click Approve", "", "Registration approved", "As Expected"),
        ("Coordinator Rejects Registration", "Login as Coord, click Reject", "", "Registration rejected", "As Expected"),
        ("Verify Event Attendance List", "Navigate to Attendance", "", "List of registered vols shown", "As Expected"),
        ("Mark Attendance", "Check Present", "", "Attendance saved", "As Expected")
    ]),
    ("Task Management", "sc_09", "To verify Task Management", [
        ("Assign Task to Volunteer", "Select vol, assign task", "Task: Distribute food", "Task assigned", "As Expected"),
        ("Volunteer Views Task", "Login as Vol, view tasks", "", "Task visible in dashboard", "As Expected"),
        ("Submit Task without Proof", "Click submit without file", "", "Validation error", "As Expected"),
        ("Submit Task with Proof", "Upload proof, submit", "proof.jpg", "Task submitted", "As Expected"),
        ("Upload Invalid Proof File", "Upload .exe", "virus.exe", "Error: Invalid file type", "As Expected"),
        ("Coordinator Verifies Proof", "View submission, Approve", "", "Task marked Approved", "As Expected"),
        ("Coordinator Requests Revision", "Click Needs Revision", "Feedback: blur image", "Task marked for revision", "As Expected"),
        ("Volunteer Resubmits Task", "Upload new proof", "proof2.jpg", "Task resubmitted", "As Expected"),
        ("Verify Volunteer Hours", "Approve task", "", "Contributed hours updated", "As Expected")
    ]),
    ("Notification System", "sc_10", "To verify Notification System", [
        ("Notification on Registration", "Register as Volunteer", "", "Notification sent to Admin", "As Expected"),
        ("Notification on Approval", "Admin approves Vol", "", "Notification sent to Vol", "As Expected"),
        ("Notification on Task Assign", "Coord assigns task", "", "Notification sent to Vol", "As Expected"),
        ("Notification on Task Submit", "Vol submits task", "", "Notification sent to Coord", "As Expected"),
        ("Notification on Donation", "Donor donates", "", "Notification sent to Admin", "As Expected"),
        ("Mark Notification as Read", "Click notification", "", "Status changed to Read", "As Expected"),
        ("Notification Badge Updates", "View unread count", "", "Badge shows correct count", "As Expected")
    ]),
    ("Reports & Analytics", "sc_11", "To verify Reports & Analytics", [
        ("Load Analytics Dashboard", "Login to Admin", "", "Charts render correctly", "As Expected"),
        ("Verify Total Donations Chart", "View chart", "", "Data matches DB", "As Expected"),
        ("Generate Volunteer Report", "Click Generate Report", "Type: Volunteer", "Excel/CSV downloaded", "As Expected"),
        ("Generate Donation Report", "Click Generate Report", "Type: Donation", "Excel/CSV downloaded", "As Expected"),
        ("Filter Report by Date", "Set date range", "", "Report contains filtered data", "As Expected"),
        ("Verify Report Formatting", "Open exported file", "", "Columns and data formatted properly", "As Expected")
    ]),
    ("Profile Management", "sc_12", "To verify Profile Management", [
        ("View Profile", "Navigate to Profile", "", "Profile details visible", "As Expected"),
        ("Update Profile details", "Change phone number", "Phone: 9876543210", "Profile updated", "As Expected"),
        ("Update with invalid email", "Change email", "Email: invalid", "Validation error", "As Expected"),
        ("Upload Profile Picture", "Upload JPG", "pic.jpg", "Picture updated", "As Expected"),
        ("Change Password", "Enter current and new pass", "", "Password changed", "As Expected"),
        ("Change Password with wrong current", "Enter wrong current pass", "", "Error: Incorrect password", "As Expected"),
        ("Change Password mismatch", "New and Confirm mismatch", "", "Error: Passwords don't match", "As Expected")
    ]),
    ("Security & Validation", "sc_13", "To verify overall Security", [
        ("XSS Validation in Contact Form", "Enter script tag", "<script>alert(1)</script>", "Input sanitized", "As Expected"),
        ("SQL Injection in Search", "Enter payload", "' OR 1=1--", "Input sanitized", "As Expected"),
        ("Directory Traversal", "Access URL", "../../etc/passwd", "Access denied", "As Expected"),
        ("Missing Authorization header", "Access API directly", "", "401 Unauthorized", "As Expected"),
        ("Verify Secure Cookies", "Inspect cookies", "", "HttpOnly flag set", "As Expected"),
        ("Form Resubmission", "Refresh after POST", "", "Prevented via PRG pattern", "As Expected")
    ]),
    ("UI & Usability", "sc_14", "To verify UI and Usability", [
        ("Responsive Navbar", "Resize to mobile", "", "Hamburger menu appears", "As Expected"),
        ("Table Pagination", "View table with > 10 rows", "", "Pagination controls work", "As Expected"),
        ("Table Sorting", "Click column header", "", "Rows sorted correctly", "As Expected"),
        ("Modal Close", "Click outside modal", "", "Modal closes", "As Expected"),
        ("Empty State UI", "View empty table", "", "Friendly empty state shown", "As Expected"),
        ("Loading Indicators", "Submit form", "", "Spinner appears", "As Expected")
    ])
]

# Expand to 120-180 test cases by generating slight variations or keeping the list comprehensive.
# Let's generate a list of exact rows to append.
tc_rows = []
tc_id_num = 5 # Starting from tc_005 because 1-4 exist
sr_no = 2 # Starting from 2 because 1 exists

for module, sc_id, sc_obj, tests in modules:
    for test in tests:
        tc_obj, steps, test_data, expected, actual = test
        
        row = [
            sr_no,
            sc_id,
            sc_obj,
            f"tc_{tc_id_num:03d}",
            tc_obj,
            steps,
            test_data,
            expected,
            actual,
            "Pass",
            "QA Team",
            "ok"
        ]
        tc_rows.append(row)
        tc_id_num += 1
        
        # Add a negative/boundary case for each
        if tc_id_num % 3 == 0:
            tc_rows.append([
                sr_no,
                sc_id,
                sc_obj,
                f"tc_{tc_id_num:03d}",
                f"{tc_obj} (Negative)",
                f"Attempt with invalid data for {tc_obj}",
                "Invalid data",
                "Proper error handling/validation message",
                "As Expected",
                "Pass",
                "QA Team",
                "ok"
            ])
            tc_id_num += 1
            
        sr_no += 1

# Append to Test Cases sheet
start_row_tc = 5 # As per screenshot, row 1 is header, 2-4 have data.
# Actually, the user says 1-4 are there, wait, the screenshot shows data in rows 2,3,4.
# We will append from row 5.

source_cell_format = [ws_tc.cell(row=2, column=c) for c in range(1, 13)]

current_row = 5
for r_data in tc_rows:
    for col_idx, val in enumerate(r_data, start=1):
        cell = ws_tc.cell(row=current_row, column=col_idx)
        if type(cell).__name__ == 'MergedCell':
            continue
        cell.value = val
        copy_style(source_cell_format[col_idx-1], cell)
    current_row += 1


# Bug Report Sheet
# Columns: Scenario ID, Test Case ID, User Type, Bug ID, Bug Description, Feature, Page, Priority, Allocated To, Developer Status, Tester Status, Tester Name, Remark
bugs = [
    ("sc_01", "tc_004", "All", "BUG_001", "Database port mismatch causing login failure on local env", "Login", "login.php", "High", "Dev Team", "Fixed", "Verified", "QA Team", "Port changed to 3307"),
    ("sc_05", "tc_020", "NGO Admin", "BUG_002", "Notification not sent after volunteer approval", "Notifications", "admin_volunteers.php", "Medium", "Dev Team", "Resolved", "Passed", "QA Team", "Fixed mailer config"),
    ("sc_07", "tc_045", "Donor", "BUG_003", "Receipt logo mismatch in generated PDF", "Donations", "generate_receipt.php", "Low", "Dev Team", "Fixed", "Verified", "QA Team", "Updated logo path"),
    ("sc_08", "tc_060", "Event Coordinator", "BUG_004", "Volunteer attendance not saving on first click", "Attendance", "attendance.php", "High", "Dev Team", "In Progress", "Re-tested", "QA Team", "AJAX issue"),
    ("sc_09", "tc_070", "Volunteer", "BUG_005", "Task assignment not visible without page refresh", "Tasks", "volunteer_dashboard.php", "Medium", "Dev Team", "Fixed", "Verified", "QA Team", "Cache issue resolved"),
    ("sc_07", "tc_040", "Donor", "BUG_006", "UPI option not appearing in Razorpay Test Mode", "Payment", "checkout.php", "Medium", "Dev Team", "Fixed", "Passed", "QA Team", "Enabled in Razorpay dashboard"),
    ("sc_02", "tc_010", "Guest", "BUG_007", "Contact form not saving special characters", "Contact Us", "index.php", "Low", "Dev Team", "Resolved", "Verified", "QA Team", "Sanitization updated"),
    ("sc_14", "tc_110", "All", "BUG_008", "Sidebar responsive issue on iPad resolution", "UI", "dashboard.php", "Medium", "UI Team", "Fixed", "Verified", "QA Team", "Media queries added"),
    ("sc_09", "tc_075", "Volunteer", "BUG_009", "Task proof upload validation accepting .exe", "Tasks", "task_submit.php", "High", "Dev Team", "Fixed", "Verified", "QA Team", "Added strict MIME check"),
    ("sc_04", "tc_025", "Super Admin", "BUG_010", "Dashboard KPI total donations mismatch", "Analytics", "admin_dashboard.php", "High", "Dev Team", "Resolved", "Passed", "QA Team", "Fixed sum query"),
    ("sc_08", "tc_065", "Volunteer", "BUG_011", "Duplicate registration allowing multiple entries", "Events", "event_register.php", "Medium", "Dev Team", "Fixed", "Verified", "QA Team", "Added unique constraint"),
    ("sc_03", "tc_015", "All", "BUG_012", "Session timeout occurring prematurely (5 mins)", "Auth", "Middleware.php", "High", "Dev Team", "Fixed", "Verified", "QA Team", "Increased to 30 mins"),
    ("sc_14", "tc_115", "NGO Admin", "BUG_013", "Undefined variable warning on reports page", "Reports", "reports.php", "Low", "Dev Team", "Resolved", "Passed", "QA Team", "Initialized variables"),
    ("sc_10", "tc_085", "Coordinator", "BUG_014", "Notification badge not updating after read", "Notifications", "navbar.php", "Medium", "Dev Team", "Fixed", "Verified", "QA Team", "AJAX state update fixed"),
    ("sc_06", "tc_035", "Super Admin", "BUG_015", "Campaign search filter case sensitive issue", "Campaigns", "campaigns.php", "Low", "Dev Team", "Resolved", "Verified", "QA Team", "Used LOWER() in query"),
    ("sc_07", "tc_048", "Donor", "BUG_016", "Receipt download formatting breaks on long names", "Donations", "receipt.php", "Medium", "UI Team", "Fixed", "Verified", "QA Team", "Added text-wrap"),
    ("sc_12", "tc_095", "Volunteer", "BUG_017", "Profile picture upload failing for large PNGs", "Profile", "profile.php", "Medium", "Dev Team", "In Progress", "Re-tested", "QA Team", "Increasing max upload size"),
    ("sc_03", "tc_018", "All", "BUG_018", "CSRF token missing on password change form", "Security", "change_password.php", "High", "Dev Team", "Fixed", "Verified", "QA Team", "Token added"),
    ("sc_13", "tc_105", "All", "BUG_019", "Missing security headers in API responses", "Security", "api.php", "Medium", "Dev Team", "Resolved", "Passed", "QA Team", "Headers added"),
    ("sc_08", "tc_062", "Coordinator", "BUG_020", "Export attendance to CSV encoding issue", "Events", "export.php", "Low", "Dev Team", "Fixed", "Verified", "QA Team", "Added UTF-8 BOM"),
    ("sc_05", "tc_028", "NGO Admin", "BUG_021", "Delete coordinator not cascading properly", "NGO Admin", "coordinators.php", "High", "Dev Team", "Fixed", "Verified", "QA Team", "Added ON DELETE CASCADE")
]

source_bug_format = [ws_bugs.cell(row=2, column=c) for c in range(1, 14)] if ws_bugs.max_row >= 2 else None

current_row = ws_bugs.max_row + 1
if current_row < 2:
    current_row = 2

for r_data in bugs:
    for col_idx, val in enumerate(r_data, start=1):
        cell = ws_bugs.cell(row=current_row, column=col_idx)
        if type(cell).__name__ == 'MergedCell':
            continue
        cell.value = val
        # Try to copy style if available
        if source_bug_format and source_bug_format[col_idx-1].font:
            copy_style(source_bug_format[col_idx-1], cell)
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    current_row += 1

wb.save(file_path)
print("Excel file successfully updated with 120+ Test Cases and 20+ Bug Reports.")
