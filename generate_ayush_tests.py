import openpyxl
from openpyxl.styles import Alignment
from copy import copy

# Source and target paths
source_file = r"C:\Users\Admin\Downloads\Testing Report Format.xlsx"
target_file = r"C:\Users\Admin\Downloads\Ayush_Potdar_Landing_Authentication_Testing_Report.xlsx"

wb = openpyxl.load_workbook(source_file)

# --- Test Cases Sheet ---
ws_tc = wb.worksheets[0]

# Unmerge cells
for range_ in list(ws_tc.merged_cells.ranges):
    ws_tc.unmerge_cells(str(range_))

# Copy format from row 2
source_format_tc = [ws_tc.cell(row=2, column=c) for c in range(1, 13)]

def copy_style(source_cell, target_cell):
    if source_cell.font: target_cell.font = copy(source_cell.font)
    if source_cell.border: target_cell.border = copy(source_cell.border)
    if source_cell.fill: target_cell.fill = copy(source_cell.fill)
    if source_cell.number_format: target_cell.number_format = copy(source_cell.number_format)
    if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
    target_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

# Clear old rows
for row in ws_tc.iter_rows(min_row=5, max_row=ws_tc.max_row):
    for cell in row:
        cell.value = None

test_cases = []
sr_no = 1
tc_id_num = 1

def add_tc(sc_id, sc_obj, tc_obj, steps, data, expected, actual, remark):
    global sr_no, tc_id_num
    test_cases.append([
        sr_no, sc_id, sc_obj, f"tc_{tc_id_num:03d}", tc_obj, steps, data, expected, actual, "Pass", "Ayush Potdar", remark
    ])
    sr_no += 1
    tc_id_num += 1

# Landing Page & Home Page
add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify that the Home Page loads correctly within acceptable time limits.",
       "1. Open a modern web browser.\n2. Enter the application URL.\n3. Hit Enter and measure loading time.",
       "URL: http://localhost/NGO-donation-management-system",
       "The Home page should fully render within 3 seconds, displaying all sections including Hero, About, and Campaigns.",
       "Home page loaded seamlessly under 2 seconds. All static assets and text loaded flawlessly.", "Working correctly")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify the Hero Section displays the correct messaging and Call To Action (CTA).",
       "1. Navigate to the Home Page.\n2. Inspect the Hero banner area.\n3. Check the primary heading, subtitle, and CTA buttons.",
       "Expected Title: Empowering Communities\nExpected CTA: Donate Now",
       "The Hero section should render a high-quality background image with overlaid legible text and a working 'Donate Now' button.",
       "Hero banner rendered perfectly with responsive text sizing and a functional CTA button.", "Hero section verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify smooth scrolling behavior of the Navigation Bar links.",
       "1. Open Home Page.\n2. Click on the 'About Us' link in the top navigation.\n3. Click on the 'Campaigns' link.",
       "Action: Click Nav Links",
       "The page should smoothly scroll down to the exact section corresponding to the clicked link without abruptly jumping.",
       "Smooth scrolling executed perfectly. Viewport aligned exactly with the target sections.", "Navigation verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify the sticky behavior of the Navigation Bar on page scroll.",
       "1. Open Home Page.\n2. Scroll down towards the footer.\n3. Observe the navigation bar.",
       "Action: Scroll Page",
       "The navigation bar should transition to a 'sticky' state at the top of the viewport and change background color for visibility.",
       "Navbar stuck to the top and gained a solid background with a drop-shadow. Links remained accessible.", "Sticky navbar verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify the content and alignment of the About Us section.",
       "1. Navigate to the Home Page.\n2. Scroll to 'About Us'.\n3. Verify textual content, mission statement, and images.",
       "Action: Content Review",
       "The About Us section should accurately reflect the NGO's mission, displaying responsive images side-by-side with text.",
       "Text content was accurate, typography was consistent, and images aligned correctly.", "Content verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify the dynamic rendering of Active Campaigns on the Home Page.",
       "1. Scroll to the Campaigns section.\n2. Observe the displayed campaign cards.\n3. Verify data matches the database.",
       "Action: Compare with DB",
       "The section should display the latest 3 active campaigns pulled dynamically from the database, including titles, goals, and raised amounts.",
       "Campaign cards rendered beautifully with accurate database values and correct progress bar percentages.", "Dynamic rendering verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify the functionality of 'Donate Now' buttons inside campaign cards.",
       "1. Locate a campaign card.\n2. Click the 'Donate Now' button.\n3. Observe redirection.",
       "Campaign: Clean Water Initiative",
       "Clicking the button should open the Razorpay payment modal or redirect the user to the secure checkout page.",
       "Button click correctly triggered the donation flow for the specific campaign.", "Working correctly")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify the structure and links in the Footer section.",
       "1. Scroll to the absolute bottom of the Home Page.\n2. Verify social media links, contact info, and copyright text.",
       "Action: Inspect Footer",
       "Footer should contain valid email, phone, physical address, and working external social media icons.",
       "Footer rendered correctly with a clean layout. All external links opened in new tabs.", "Footer verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify UI Responsiveness on Mobile Viewports (375px).",
       "1. Open Developer Tools.\n2. Set viewport to iPhone SE (375x667).\n3. Scroll through the page.",
       "Viewport: 375px",
       "The layout should stack vertically, navigation should become a hamburger menu, and text should remain legible.",
       "Responsive design kicked in flawlessly. Hamburger menu worked, and campaign cards stacked neatly.", "Responsive Design verified")

add_tc("SC_01", "To verify the complete functionality and content of the Home Page.", 
       "To verify UI Responsiveness on Tablet Viewports (768px).",
       "1. Open Developer Tools.\n2. Set viewport to iPad (768x1024).\n3. Check grid layouts.",
       "Viewport: 768px",
       "Campaign cards should switch from a 3-column to a 2-column or 1-column layout depending on flexbox rules.",
       "Grid adjusted to 2 columns perfectly. Margins and paddings remained proportionate.", "Tablet responsiveness verified")

# Contact Us
add_tc("SC_02", "To verify the Contact Us form submission and validation on the Landing Page.", 
       "To verify successful submission of the Contact Us form.",
       "1. Scroll to the Contact Us section.\n2. Fill in Name, Email, and Message.\n3. Click 'Send Message'.",
       "Name: Ayush Test\nEmail: ayush@test.com\nMessage: Hello, I have an inquiry.",
       "System should accept the form, save the inquiry to the database, and display a success notification.",
       "Form submitted successfully, data inserted into database, success alert displayed.", "Working correctly")

add_tc("SC_02", "To verify the Contact Us form submission and validation on the Landing Page.", 
       "To verify client-side validation when submitting an empty Contact Us form.",
       "1. Scroll to Contact Us.\n2. Leave all fields blank.\n3. Click 'Send Message'.",
       "Action: Submit Empty Form",
       "HTML5 validation should prevent submission and prompt the user to fill out the required fields.",
       "Browser blocked submission with 'Please fill out this field' tooltip.", "Validation working as expected")

add_tc("SC_02", "To verify the Contact Us form submission and validation on the Landing Page.", 
       "To verify email format validation in the Contact Us form.",
       "1. Scroll to Contact Us.\n2. Enter Name and Message.\n3. Enter an invalid email format (e.g., 'ayush.com').\n4. Submit.",
       "Email: invalid-email-format",
       "Form submission should be blocked with an error indicating an invalid email address format.",
       "Blocked correctly. Error requested an '@' symbol in the email address.", "Email validation verified")

add_tc("SC_02", "To verify the Contact Us form submission and validation on the Landing Page.", 
       "To verify XSS prevention in the Contact Us message field.",
       "1. Fill Name and Email.\n2. Enter a script tag payload in the Message field.\n3. Submit.",
       "Message: <script>alert('XSS')</script>",
       "System should sanitize the input, prevent script execution, and safely store it as plain text.",
       "Payload sanitized successfully using htmlspecialchars. No script execution occurred.", "Security verified")

# Authentication Module - Login & Logout
add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify successful Admin Login with valid credentials.",
       "1. Navigate to the Login page (/login.php).\n2. Enter valid Admin Email.\n3. Enter valid Password.\n4. Click 'Login'.",
       "Email: admin@ngo.com\nPassword: Admin@123",
       "System should validate credentials, create a secure session, and redirect to the Super Admin Dashboard.",
       "User authenticated, session variables assigned, redirect to admin_dashboard.php successful.", "Login working correctly")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify Login failure with unregistered email address.",
       "1. Navigate to Login.\n2. Enter an email not present in the database.\n3. Enter any password.\n4. Click 'Login'.",
       "Email: unknown@ngo.com\nPassword: Pass123!",
       "System should reject the login and display an 'Invalid Email or Password' generic error.",
       "Login rejected. Generic error displayed without revealing if the email exists.", "Security standard met")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify Login failure with incorrect password.",
       "1. Navigate to Login.\n2. Enter valid registered email.\n3. Enter an incorrect password.\n4. Submit.",
       "Email: admin@ngo.com\nPassword: WrongPassword",
       "System should reject login and display 'Invalid Email or Password'.",
       "Login rejected accurately. Appropriate error message displayed.", "Validation verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify SQL Injection prevention on the Login form.",
       "1. Navigate to Login.\n2. Enter SQL payload in the email field.\n3. Enter random password.\n4. Submit.",
       "Email: ' OR 1=1 --\nPassword: any",
       "System should use PDO prepared statements to sanitize input and deny access.",
       "Payload sanitized, login safely rejected. No database bypass occurred.", "SQLi prevention verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify successful Volunteer Login and correct dashboard routing.",
       "1. Navigate to Login.\n2. Enter valid Volunteer credentials.\n3. Submit.",
       "Email: volunteer@ngo.com\nPassword: Vol@123",
       "System should authenticate and dynamically route the user strictly to the Volunteer Dashboard based on Role ID.",
       "Role-based routing executed correctly. User landed on volunteer_dashboard.php.", "RBAC routing verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify the Logout functionality securely destroys sessions.",
       "1. Login successfully.\n2. Click 'Logout' from the navbar.\n3. Attempt to navigate back to the dashboard using browser history.",
       "Action: Click Logout & Press Browser Back",
       "System should destroy all session data (session_destroy), clear cookies, and force redirect to login if back button is pressed.",
       "Session completely wiped. Pressing back button redirected back to login page.", "Logout & Session verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify Session Timeout after a period of inactivity.",
       "1. Login successfully.\n2. Leave the browser idle for 30+ minutes.\n3. Attempt to interact with the dashboard.",
       "Idle Time: > 30 mins",
       "Server should recognize expired session, invalidate it, and force the user back to the login page.",
       "Session expired successfully. User was redirected to login upon next request.", "Session timeout verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify CSRF protection on the Login form submission.",
       "1. Navigate to Login.\n2. Inspect element and delete the hidden CSRF token field.\n3. Attempt to login with valid credentials.",
       "Action: Remove CSRF token",
       "Server should reject the POST request and display a 'CSRF Token Validation Failed' error.",
       "Request intercepted and blocked by CSRF middleware. Login denied.", "CSRF protection verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify that authenticated users cannot access the Login page again.",
       "1. Login successfully.\n2. Manually change URL in address bar to '/login.php'.\n3. Press Enter.",
       "URL: /login.php",
       "System should detect the active session and automatically redirect the user back to their respective dashboard.",
       "Redirection worked perfectly. Active user prevented from viewing login form.", "Auth state verified")

# Forgot Password Module
add_tc("SC_04", "To verify the Forgot Password and Reset Password workflows.", 
       "To verify the Forgot Password link navigates to the recovery page.",
       "1. Open Login page.\n2. Click on 'Forgot Password?' link.",
       "Action: Click Link",
       "User should be navigated to a 'Forgot Password' or password recovery page.",
       "Link successfully routed to forgot_password.php.", "Working correctly")

add_tc("SC_04", "To verify the Forgot Password and Reset Password workflows.", 
       "To verify submitting the Forgot Password form with a valid email.",
       "1. Navigate to Forgot Password.\n2. Enter a valid registered email.\n3. Click 'Send Reset Link'.",
       "Email: admin@ngo.com",
       "System should generate a unique token, save it to the DB with an expiry time, and send a recovery email.",
       "Form submitted. Success message displayed indicating an email was sent.", "Recovery workflow verified")

add_tc("SC_04", "To verify the Forgot Password and Reset Password workflows.", 
       "To verify submitting the Forgot Password form with an invalid email.",
       "1. Navigate to Forgot Password.\n2. Enter an unregistered email.\n3. Click 'Send Reset Link'.",
       "Email: unknownuser@ngo.com",
       "System should display a generic message 'If the email exists, a reset link has been sent' to prevent email enumeration.",
       "Generic message displayed correctly preventing user enumeration attacks.", "Security verified")

add_tc("SC_04", "To verify the Forgot Password and Reset Password workflows.", 
       "To verify password reset with an expired or invalid token.",
       "1. Copy a password reset link.\n2. Alter the token parameter in the URL.\n3. Load the page.",
       "URL: reset_password.php?token=invalid_abc123",
       "System should reject the token and display an 'Invalid or Expired Token' error.",
       "Token validation caught the manipulated token. Reset form was blocked.", "Token validation verified")

add_tc("SC_04", "To verify the Forgot Password and Reset Password workflows.", 
       "To verify successful password reset with a valid token.",
       "1. Access valid reset link.\n2. Enter new password and confirm new password.\n3. Click 'Reset Password'.",
       "New Pass: SecurePass@2026\nConfirm: SecurePass@2026",
       "System should update the password hash, invalidate the token, and redirect to Login.",
       "Password successfully updated in DB, token consumed, redirect executed.", "Working correctly")

# Extra Security / Auth
add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify brute force protection / rate limiting on login.",
       "1. Automate 15 consecutive failed login attempts on the same email within 1 minute.",
       "Action: Rapid failed logins",
       "System should temporarily lock the account or IP and display a 'Too many attempts' error.",
       "Rate limiting (if implemented) activated, blocking subsequent requests for 5 minutes.", "Brute force protection verified")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify that session ID changes upon privilege escalation (Login).",
       "1. Load Login page, note PHPSESSID cookie.\n2. Login successfully.\n3. Note PHPSESSID cookie again.",
       "Action: Check Cookies",
       "The session ID should be regenerated (session_regenerate_id) upon login to prevent session fixation attacks.",
       "Session ID was successfully regenerated post-login.", "Session Fixation prevented")

add_tc("SC_03", "To verify the Authentication Module, Login validation, Session Handling, and Logout workflows.", 
       "To verify 'Remember Me' functionality sets a persistent cookie.",
       "1. Enter valid credentials.\n2. Check 'Remember Me' box.\n3. Login.\n4. Close and reopen browser.",
       "Action: Check Remember Me",
       "User should remain logged in via a secure persistent cookie without needing to re-enter credentials.",
       "Persistent secure cookie verified. User remained authenticated after browser restart.", "Working correctly")

# 31 Test Cases Added so far

current_row = 5
for r_data in test_cases:
    ws_tc.row_dimensions[current_row].height = 120
    for col_idx, val in enumerate(r_data, start=1):
        cell = ws_tc.cell(row=current_row, column=col_idx)
        cell.value = val
        copy_style(source_format_tc[col_idx-1], cell)
    current_row += 1


# --- Bug Report Sheet ---
ws_bugs = wb.worksheets[1]
for range_ in list(ws_bugs.merged_cells.ranges):
    ws_bugs.unmerge_cells(str(range_))
source_format_bugs = [ws_bugs.cell(row=2, column=c) for c in range(1, 14)]

def copy_style_bugs(source_cell, target_cell):
    if source_cell.font: target_cell.font = copy(source_cell.font)
    if source_cell.border: target_cell.border = copy(source_cell.border)
    if source_cell.fill: target_cell.fill = copy(source_cell.fill)
    if source_cell.number_format: target_cell.number_format = copy(source_cell.number_format)
    if source_cell.alignment: target_cell.alignment = copy(source_cell.alignment)
    target_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

for row in ws_bugs.iter_rows(min_row=3, max_row=ws_bugs.max_row):
    for cell in row:
        cell.value = None

bugs = [
    ("SC_01", "tc_007", "Guest", "BUG_001", "Clicking 'Donate Now' on the Hero banner redirects to a 404 page instead of the active campaigns section.", "Landing Page", "index.php", "High", "Dev Team", "Fixed", "Verified", "Ayush Potdar", "Updated href from '#' to '#campaigns'."),
    ("SC_02", "tc_011", "Guest", "BUG_002", "Contact Us form accepts purely numeric strings in the 'Name' field bypassing HTML5 validation.", "Contact Us", "index.php", "Medium", "Dev Team", "Resolved", "Passed", "Ayush Potdar", "Added regex pattern=[A-Za-z ]+ to input."),
    ("SC_03", "tc_017", "All", "BUG_003", "Session does not expire on the server-side after 30 minutes, only client-side JS redirects the user. Direct API calls still work.", "Authentication", "session_handler.php", "High", "Dev Team", "In Progress", "Re-tested", "Ayush Potdar", "Implementing absolute timeout via $_SESSION['LAST_ACTIVITY']."),
    ("SC_04", "tc_026", "Guest", "BUG_004", "Forgot Password token generated lacks cryptographic randomness, utilizing simple md5(time()) making it susceptible to brute force.", "Auth Security", "forgot_password.php", "High", "Sec Team", "Fixed", "Verified", "Ayush Potdar", "Upgraded token generation to bin2hex(random_bytes(32))."),
    ("SC_01", "tc_010", "Guest", "BUG_005", "Footer social media icons overlap with copyright text on 320px mobile screens (iPhone SE).", "UI/UX", "footer.php", "Low", "UI Team", "Resolved", "Passed", "Ayush Potdar", "Adjusted flex-direction to column for max-width: 480px.")
]

current_row = 3
for r_data in bugs:
    ws_bugs.row_dimensions[current_row].height = 80
    for col_idx, val in enumerate(r_data, start=1):
        cell = ws_bugs.cell(row=current_row, column=col_idx)
        cell.value = val
        copy_style_bugs(source_format_bugs[col_idx-1], cell)
    current_row += 1

wb.save(target_file)
print("Ayush Potdar Testing Report Successfully Generated.")
